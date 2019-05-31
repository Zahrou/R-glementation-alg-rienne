# -*- coding: utf-8 -*-
"""
Created on Sat Oct 27 22:34:01 2018

@author: Zahreddine
"""

# -*- coding: utf-8 -*-
"""
Created on Sat Jun  9 01:17:25 2018

@author: Toshiba
"""

import openpyxl
import os 
import re



os.chdir("D:\\Desktop\\Journal officiel2")



#Create a new workbook  



wb = openpyxl.load_workbook("database3.xlsx")
   
def populate_decret(text):
    
    worksheet = wb.get_sheet_by_name(text)
    texts =  []
   
               
      
    #print(worksheet.max_row)    
    count = 0        
    for r in range(2, worksheet.max_row+1):
            
        T = worksheet.cell(row=r, column=1).value
        #print(T)
        try:        
            num = re.search(r'[0-9]{2}-[0-9]{2}([0-9]{1})?', T).group(0)
        except AttributeError:
            num = ' '
        try:
            date = re.search(r'[0-9]{1}((er)|([0-9]{1}))?(\s)(janvier|février|mars|avril|mai|juin|juillet|août|aôut|aout|septembre|octobre|novembre|décembre)(\s)*(1|2)[0-9]{3}', T, re.IGNORECASE).group(0)
        except AttributeError:
            date = re.search(r'20[0-9]{2}', T).group(0)
        
        objet = T[T.index(re.search(r'(19|20)[0-9]{2}', T).group(0))+5:]
        urljo = worksheet.cell(row=r, column=4).value
        text_attributes = (num, date, objet, urljo)
        texts.append(text_attributes)
        count += 1
        print(count)
    return  texts
            
        

def clean_data(text):
    worksheet = wb.get_sheet_by_name(text)
    count = 0
    for r in range(2, worksheet.max_row+1):
        T = worksheet.cell(row=r, column=1).value
        print("stopped in line 66")
        print("\n")
        try:
            T = T[:T.index(".")]
        except ValueError:
            try:
                T = T[:T.index("JOURNAL")] or T[:T.index("S O M M")] or T[:T.index("SOMM")]
                print("stopped in line 72")
            except:
                pass
        T = T.replace("™", "'").replace("‹", "à")
        print("stopped in line 74")
        to_be_changed = None
        try:
            to_be_changed = re.search(r'(D|d|L|l|)?(ê)(a|é|e|o|i|u|h|A|E|O|I|U|H)', T).group(0)
        except :
            print("stopped in line 79")
            pass
        if to_be_changed:
            new = to_be_changed.replace("ê", "'")
            print("stopped in line 82")
            T = T.replace(to_be_changed, new)
            print("stopped in line 85")
            
        T = T + "."
        print("stopped in line 87")
        worksheet.cell(row=r, column=1).value = T
        count += 1
    print(count)
                      
    wb.save("database3.xlsx")
        
        
    


def populate_journal():
    
    
    NumeroDate = []
    URLs =  []
               
      
    #print(worksheet.max_row)    
    for sheet in wb.get_sheet_names():
        worksheet = wb.get_sheet_by_name(sheet)        
        for r in range(2, worksheet.max_row+1):
            
            NumDate = worksheet.cell(row=r, column=2).value+"du"+worksheet.cell(row=r, column=3).value
                
            if NumDate not in  NumeroDate:
                NumeroDate.append(NumDate)
                
            url = worksheet.cell(row=r, column=4).value
            if url not in  URLs:
                URLs.append(url)
            
    return NumeroDate, URLs