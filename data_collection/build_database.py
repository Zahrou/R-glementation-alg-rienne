# -*- coding: utf-8 -*-
"""
Created on Wed Nov 29 22:21:18 2017

@author: Toshiba
"""
import os
import sys
import openpyxl
sys.path.append("D:/Desktop/JORADP")


from TextExtractor import journalOfficiel, get_string


parent = "D:\\Desktop\\Journal officiel2"
os.chdir(parent)
#journaux_2002 = open("2002_1.txt", "w")
wb = openpyxl.Workbook()

NonRetrievedJournals = []

for year in os.listdir():
    if os.path.isdir(os.path.join(parent, year)):
        directory = os.path.join(parent, year)
        os.chdir(directory)
        print("retrieving data from %s" % year)
        
        for file in os.listdir():
            print(("openning journal : %s") % file[:-4])
            if  file[-4:] == '.pdf':
                p = journalOfficiel(file, get_string(file)[0])
                p.construir_journal()
                
                for text in p.texts:
                    if len(p.texts[text]) != 0:
                        
                        if text not in wb.get_sheet_names():
                            wb.create_sheet(text)
                            worksheet = wb.get_sheet_by_name(text)
                            
                            worksheet["A1"] = "Objet"
                            worksheet["B1"] = "Numero jo"
                            worksheet["C1"] = "Date jo"
                            worksheet["D1"] = "URL jo"
                        else:
                            worksheet = wb.get_sheet_by_name(text)
                        print("adding dat to %s"% worksheet)
                        Lines = []
                        for e in p.texts[text]:
                            #print(text)
                            
                            line = []
                            line.append(e)
                            line.append(p.num_journal)
                            line.append(p.date_journal)
                            line.append(p.journal_url)
                            Lines.append(line)
                            
                        for row in Lines:
                            #print("\t \t", " ", row)
                            worksheet.append(row)
                    
            
            Lines = []        
    
            
                
os.chdir(parent)
wb.save("database3.xlsx")



#p = journalOfficiel("Jouranl 2010072.pdf", get_string("Jouranl 2010072.pdf")[0])
#p.construir_journal()
#        
#for text in p.texts:
#    if len(p.texts[text]) != 0:
#        
#        if text not in wb.get_sheet_names():
#            wb.create_sheet(text)
#            worksheet = wb.get_sheet_by_name(text)
#            
#            worksheet["A1"] = "Objet"
#            worksheet["B1"] = "Numero jo"
#            worksheet["C1"] = "Date jo"
#            worksheet["D1"] = "URL jo"
#        else:
#            worksheet = wb.get_sheet_by_name(text)
#        print(worksheet)
#        Lines = []
#        for e in p.texts[text]:
#            print(text)
#            
#            line = []
#            line.append(e)
#            line.append(p.num_journal)
#            line.append(p.date_journal)
#            line.append(p.journal_url)
#            Lines.append(line)
#            
#        for row in Lines:
#            print("\t \t", " ", row)
#            worksheet.append(row)
##Lines = []        
#    
#            
#                
#
#wb.save("database.xlsx")
        











